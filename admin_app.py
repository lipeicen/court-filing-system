#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
法院保全案件后台管理系统
功能：案件数据录入、财产线索管理、文件上传、案件列表查看、Excel批量导入
"""

import os
import sys
import shutil
import pymysql
import io
import re
import time
from datetime import datetime
from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file, session
from werkzeug.utils import secure_filename
import logging

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(r'/app/admin_app.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

app = Flask(__name__, template_folder='admin/templates', static_folder='admin/static')
app.secret_key = 'court_filing_secret_key_2024'

# 登录验证装饰器
def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# 自动立案任务状态
auto_filing_status = {
    'running': False,
    'total': 0,
    'current': 0,
    'current_case': None,
    'completed': [],
    'failed': [],
    'start_time': None
}

# 数据库配置
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': 'lijiayu123',
    'database': 'court_filing',
    'charset': 'utf8mb4'
}

UPLOAD_BASE = r'/app/uploads'
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'jpg', 'jpeg', 'png'}
PROPERTY_TYPES = [
    '存款', '房产', '土地使用权', '集体土地所有权', '森林、林木所有权',
    '船舶', '珠宝玉石首饰', '债权', '知识产权', '股权', '股票', '基金',
    '债券', '交通运输工具', '机器设备', '货币现金', '产品、原材料',
    '探矿、采矿权', '古玩字画', '留置的货物', '船用物料', '船用燃料',
    '船载货物', '其他'
]
GUARANTEE_TYPES = ['提供保证人', '设定抵押', '设定质押', '交纳保证金']
APPLICANT_TYPES = ['自然人', '法人', '非法人组织']
RESPONDENT_TYPES = ['自然人', '法人', '非法人组织']
FILE_CATEGORIES = ['保全申请书', '起诉状', '立案受理通知书', '身份证明材料', '担保材料', '证据材料', '设定质押', '其他材料']

# ==================== Excel模板字段定义 ====================
EXCEL_TEMPLATE_COLUMNS = [
    ('case_no', '案件编号*', True), ('case_name', '案件名称*', True), ('preserve_type', '保全类别', False),
    ('case_reason', '案由', False), ('court_name', '申请法院*', True), ('preserve_amount', '保全金额*', True),
    ('delivery_address', '送达地址', False), ('contact_name', '联系人姓名', False), ('contact_phone', '联系人电话', False),
    ('remark', '备注', False), ('applicant_type', '申请人类型', False), ('applicant_name', '申请人姓名*', True),
    ('applicant_cert_no', '申请人身份证号*', True), ('applicant_gender', '申请人性别', False),
    ('applicant_phone', '申请人手机号*', True), ('applicant_address', '申请人地址', False),
    ('applicant_uscc', '申请人统一社会信用代码', False),
    ('applicant_nature', '申请人单位性质', False),
    ('applicant_legal_person', '申请人法定代表人', False),
    ('applicant_legal_title', '申请人法定代表人职务', False),
    ('applicant_reg_address', '申请人单位注册地', False),
    ('applicant_tel', '申请人固定电话', False), ('applicant_email', '申请人邮箱', False),
    ('applicant_birth', '申请人生日', False), ('applicant_age', '申请人年龄', False),
    ('applicant_nation', '申请人民族', False), ('applicant_household_address', '申请人户籍地址', False),
    ('respondent_type', '被申请人类型', False), ('respondent_name', '被申请人姓名*', True),
    ('respondent_id', '被申请人身份证号*', True), ('respondent_gender', '被申请人性别', False),
    ('respondent_uscc', '被申请人统一社会信用代码', False),
    ('respondent_phone', '被申请人手机号', False), ('respondent_address', '被申请人地址', False),
    ('respondent_tel', '被申请人固定电话', False), ('respondent_residence', '被申请人经常居住地', False),
    ('respondent_nature', '被申请人性质', False), ('respondent_legal_person', '被申请人法定代表人', False),
    ('respondent_legal_title', '被申请人法定职务', False), ('respondent_reg_address', '被申请人户籍地址', False),
    ('property_type', '财产类型*', True), ('property_owner', '财产所有人', False),
    ('bank_name', '开户银行', False), ('bank_account', '银行账号', False), ('amount', '数额', False),
    ('currency', '币种', False), ('property_value', '财产价值', False), ('property_province', '财产所在省份', False),
    ('property_city', '财产所在城市', False), ('property_location', '财产详细地址', False),
    ('property_cert_no', '产权证号', False), ('property_detail_location', '财产具体位置', False),
    ('vehicle_plate_no', '车牌号', False), ('vehicle_location', '车辆位置', False),
    ('stock_name', '股票名称', False), ('stock_code', '股票代码', False), ('stock_quantity', '持股数量', False),
    ('stock_reg_location', '发行单位注册地', False), ('stock_company_name', '持股公司名称', False),
    ('stock_ratio', '出资比例', False), ('equipment_name', '设备名称', False),
    ('guarantee_type', '担保方式', False), ('guarantee_value', '担保价值', False),
    ('guarantor_name', '担保人姓名', False), ('guarantee_object', '担保物名称', False),
    ('guarantee_property_type', '担保财产类型', False), ('guarantee_remark', '担保备注', False),
    ('agent_type', '代理人类型', False), ('agent_name', '代理人姓名', False),
    ('agent_cert_no', '代理人身份证号', False), ('agent_phone', '代理人电话', False),
    ('agent_license_no', '代理人执业证号', False), ('agent_law_firm', '代理人律所', False),
]


def get_db():
    return pymysql.connect(**DB_CONFIG)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def save_uploaded_file(case_no, category, file, applicant_name=None, respondent_name=None):
    if file and allowed_file(file.filename):
        case_dir = os.path.join(UPLOAD_BASE, case_no, category)
        os.makedirs(case_dir, exist_ok=True)
        
        # 从原始文件名提取扩展名
        ext = os.path.splitext(file.filename)[1]
        
        # 格式：案号-被申请人姓名-类别.扩展名
        name = respondent_name or '未知'
        base = re.sub(r'[///:*?"<>|]', '_', f"{case_no}-{name}-{category}").strip(' .')
        new_filename = f"{base}{ext}"
        filepath = os.path.join(case_dir, new_filename)
        if os.path.exists(filepath):
            os.remove(filepath)
        file.save(filepath)
        return filepath
    return None


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        
        # 从 users 表验证账号密码（支持多用户）
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT username, password, user_type FROM system_users WHERE username = %s AND is_active = 1", (username,))
                user = cur.fetchone()
                
                if user and user[1] == password:
                    session['logged_in'] = True
                    session['username'] = username
                    session['user_type'] = user[2]
                    
                    # 同步到 system_config（自动立案用）
                    cur.execute("UPDATE system_config SET config_value = %s WHERE config_key = 'login_username'", (username,))
                    cur.execute("UPDATE system_config SET config_value = %s WHERE config_key = 'login_password'", (password,))
                    cur.execute("UPDATE system_config SET config_value = %s WHERE config_key = 'login_user_type'", (user[2],))
                    conn.commit()
                    
                    flash('登录成功', 'success')
                    return redirect(url_for('index'))
                else:
                    flash('账号或密码错误', 'error')
        finally:
            conn.close()
    
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        user_type = request.form.get('user_type', '个人用户').strip()
        
        if not username or not password:
            flash('账号和密码不能为空', 'error')
            return redirect(url_for('register'))
        
        if password != confirm_password:
            flash('两次输入的密码不一致', 'error')
            return redirect(url_for('register'))
        
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute('SELECT id FROM users WHERE username = %s', (username,))
                if cur.fetchone():
                    flash('该账号已被注册', 'error')
                    return redirect(url_for('register'))
                
                cur.execute('INSERT INTO system_users (username, password, user_type) VALUES (%s, %s, %s) ON DUPLICATE KEY UPDATE password = %s, user_type = %s',
                           (username, password, user_type, password, user_type))
                
                conn.commit()
                flash('注册成功，请登录', 'success')
                return redirect(url_for('login'))
        except Exception as e:
            conn.rollback()
            flash(f'注册失败: {str(e)}', 'error')
            return redirect(url_for('register'))
        finally:
            conn.close()
    
    return render_template('register.html')



@app.route('/my-filings')
@login_required
def my_filings():
    """我的立案 - 显示从法院同步的立案状态（按类别分标签页）"""
    conn = pymysql.connect(
        host="localhost", user="root", password="lijiayu123",
        database="court_filing_status", charset="utf8mb4"
    )
    try:
        with conn.cursor() as cur:
            table_map = {
                '审判': 'filing_status_trial',
                '执行': 'filing_status_execution',
                '保全': 'filing_status_preservation',
                '调解': 'filing_status_mediation',
                '破产': 'filing_status_bankruptcy',
                '信访': 'filing_status_petition'
            }
            category_counts = {}
            total_count = 0
            for cat, table in table_map.items():
                try:
                    cur.execute(f"SELECT COUNT(*) FROM {table}")
                    count = cur.fetchone()[0]
                    category_counts[cat] = count
                    total_count += count
                except:
                    category_counts[cat] = 0

            categories = {
                '审判': 'filings_trial',
                '执行': 'filings_execution',
                '保全': 'filings_preservation',
                '调解': 'filings_mediation',
                '破产': 'filings_bankruptcy',
                '信访': 'filings_petition'
            }

            filings_by_category = {}
            for cat, var_name in categories.items():
                table = table_map[cat]
                # 只取最新同步的500条（约50页），按sync_time正序
                cur.execute(f"""
                    SELECT case_no, case_name, court_name, status, status_code,
                           review_opinion, apply_date, sync_time, '{cat}' as case_category, is_new
                    FROM {table}
                    WHERE sync_time >= (
                        SELECT sync_time FROM (
                            SELECT DISTINCT sync_time 
                            FROM {table} 
                            ORDER BY sync_time DESC 
                            LIMIT 50
                        ) as t 
                        ORDER BY sync_time ASC 
                        LIMIT 1
                    )
                    ORDER BY sync_time ASC
                    LIMIT 500
                """)
                filings = []
                for row in cur.fetchall():
                    filings.append({
                        "case_no": row[0], "case_name": row[1], "court_name": row[2],
                        "status": row[3], "status_code": row[4],
                        "review_opinion": row[5], "apply_date": row[6],
                        "sync_time": row[7], "case_category": row[8],
                        "is_new": row[9]
                    })
                filings_by_category[var_name] = filings
    finally:
        conn.close()

    return render_template("my_filings.html",
                         total_count=total_count,
                         category_counts=category_counts,
                         sync_url='/sync/filing-status',
                         **filings_by_category)

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    session.pop('username', None)
    flash('已退出登录', 'success')
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    conn = get_db()
    try:
        with conn.cursor() as cur:
            # 获取当前登录用户
            current_user = session.get("username", "")
            user_type = session.get("user_type", "个人用户")
            
            # 管理员看全部，普通用户只看自己的
            if user_type == "管理员":
                cur.execute("""
                    SELECT c.id, c.case_no, c.case_name, c.applicant_name,
                           c.respondent_name, c.preserve_amount, c.court_name,
                           c.guarantee_type, c.status, c.created_at,
                           p.property_type, c.filing_status
                    FROM cases c
                    LEFT JOIN property_clues p ON c.id = p.case_id
                    ORDER BY c.created_at DESC
                """)
            else:
                cur.execute("""
                    SELECT c.id, c.case_no, c.case_name, c.applicant_name,
                           c.respondent_name, c.preserve_amount, c.court_name,
                           c.guarantee_type, c.status, c.created_at,
                           p.property_type, c.filing_status
                    FROM cases c
                    LEFT JOIN property_clues p ON c.id = p.case_id
                    WHERE c.created_by = %s
                    ORDER BY c.created_at DESC
                """, (current_user,))
            cases = []
            for row in cur.fetchall():
                cases.append({
                    "id": row[0], "case_no": row[1], "case_name": row[2],
                    "applicant_name": row[3], "respondent_name": row[4],
                    "preserve_amount": row[5], "court_name": row[6],
                    "guarantee_type": row[7], "status": row[8],
                    "created_at": row[9], "property_type": row[10],
                    "filing_status": row[11]
                })
            # 统计已自动立案数 (status=1)
            cur.execute("SELECT COUNT(*) FROM cases WHERE status = 1")
            filed_count = cur.fetchone()[0]
    finally:
        conn.close()
    # 获取系统配置
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT config_key, config_value FROM system_config")
            config_rows = cur.fetchall()
            system_config = {row[0]: row[1] for row in config_rows}
    finally:
        conn.close()
    
    # 获取当前登录用户信息
    current_user_info = {
        "login_username": session.get("username", ""),
        "login_user_type": session.get("user_type", "个人用户")
    }
    
    return render_template("index.html", cases=cases, filed_count=filed_count, has_openpyxl=HAS_OPENPYXL, system_config=system_config, current_user=current_user_info)

@login_required
@app.route('/system-config', methods=['GET', 'POST'])
def system_config_page():
    """系统配置页面：修改法院登录账号密码和身份"""
    conn = get_db()
    try:
        with conn.cursor() as cur:
            if request.method == 'POST':
                username = request.form.get('login_username', '').strip()
                password = request.form.get('login_password', '').strip()
                user_type = request.form.get('login_user_type', '个人用户').strip()
                
                if username:
                    cur.execute("UPDATE system_config SET config_value = %s WHERE config_key = 'login_username'", (username,))
                if password:
                    cur.execute("UPDATE system_config SET config_value = %s WHERE config_key = 'login_password'", (password,))
                
                # 同步更新 system_users 表
                if username and password:
                    cur.execute("INSERT INTO system_users (username, password, user_type) VALUES (%s, %s, %s) ON DUPLICATE KEY UPDATE password = %s, user_type = %s",
                               (username, password, user_type, password, user_type))
                
                conn.commit()
                flash('系统配置已更新', 'success')
                return redirect(url_for('system_config_page'))
            
            cur.execute("SELECT config_key, config_value FROM system_config")
            config_rows = cur.fetchall()
            config = {row[0]: row[1] for row in config_rows}
    finally:
        conn.close()
    
    user_types = ['个人用户', '律师用户', '法人用户']
    return render_template('system_config.html', config=config, user_types=user_types)


@login_required
@app.route('/case/<int:case_id>')
def view_case(case_id):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM cases WHERE id = %s", (case_id,))
            case_row = cur.fetchone()
            if not case_row:
                flash('案件不存在', 'error')
                return redirect(url_for('index'))
            cur.execute("SHOW COLUMNS FROM cases")
            case_cols = [c[0] for c in cur.fetchall()]
            case = dict(zip(case_cols, case_row))

            cur.execute("SELECT * FROM property_clues WHERE case_id = %s", (case_id,))
            prop_rows = cur.fetchall()
            cur.execute("SHOW COLUMNS FROM property_clues")
            prop_cols = [c[0] for c in cur.fetchall()]
            properties = [dict(zip(prop_cols, row)) for row in prop_rows]

            cur.execute("SELECT * FROM case_files WHERE case_id = %s", (case_id,))
            file_rows = cur.fetchall()
            cur.execute("SHOW COLUMNS FROM case_files")
            file_cols = [c[0] for c in cur.fetchall()]
            files = [dict(zip(file_cols, row)) for row in file_rows]
    finally:
        conn.close()
    return render_template('case_detail.html', case=case, properties=properties, files=files)


@login_required
@app.route('/case/add', methods=['GET', 'POST'])
def add_case():
    if request.method == 'POST':
        case_no = request.form.get('case_no', '').strip()
        if not case_no:
            flash('案件编号不能为空', 'error')
            return redirect(url_for('add_case'))

        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT id FROM cases WHERE case_no = %s", (case_no,))
                if cur.fetchone():
                    flash('案件编号已存在', 'error')
                    return redirect(url_for('add_case'))

                # 获取当前登录用户
                current_user = ''
                try:
                    cur.execute("SELECT config_value FROM system_config WHERE config_key = 'login_username'")
                    row = cur.fetchone()
                    if row:
                        current_user = row[0]
                except:
                    pass
                
                sql = """INSERT INTO cases (
                    case_no, case_name, preserve_type, case_reason,
                    applicant_name, applicant_id, applicant_phone, applicant_address,
                    applicant_type, applicant_cert_type, applicant_cert_no, applicant_uscc, applicant_gender,
                    applicant_country, applicant_birth, applicant_age, applicant_nation, applicant_tel,
                    applicant_reg_address, applicant_reg_country,
                    applicant_nature, applicant_legal_person, applicant_legal_title,
                    respondent_name, respondent_id, respondent_phone, respondent_address,
                    respondent_type, respondent_cert_type, respondent_uscc, respondent_gender,
                    respondent_country, respondent_birth, respondent_age, respondent_nation,
                    respondent_tel, respondent_residence,
                    respondent_nature, respondent_legal_person, respondent_legal_title,
                    respondent_reg_address, respondent_reg_country,
                    preserve_amount, court_name, guarantee_type, guarantee_value,
                    guarantor_name, pledge_property_type, mortgage_property_type,
                    guarantee_object, guarantee_remark,
                    created_by
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""

                values = (
                    case_no, request.form.get('case_name', ''),
                    request.form.get('preserve_type', '诉讼保全'),
                    request.form.get('case_reason', ''),
                    request.form.get('applicant_name', ''),
                    request.form.get('applicant_id', ''),
                    request.form.get('applicant_phone', ''),
                    request.form.get('applicant_address', ''),
                    request.form.get('applicant_type', '自然人'),
                    request.form.get('applicant_cert_type', '身份证'),
                    request.form.get('applicant_cert_no', ''),
                    request.form.get('applicant_uscc', ''),
                    request.form.get('applicant_gender', ''),
                    request.form.get('applicant_country', '中国'),
                    request.form.get('applicant_birth', ''),
                    request.form.get('applicant_age', ''),
                    request.form.get('applicant_nation', '汉族'),
                    request.form.get('applicant_tel', ''),
                    request.form.get('applicant_reg_address', ''),
                    request.form.get('applicant_reg_country', '中国'),
                    request.form.get('applicant_nature', ''),
                    request.form.get('applicant_legal_person', ''),
                    request.form.get('applicant_legal_title', ''),
                    request.form.get('respondent_name', ''),
                    request.form.get('respondent_id', ''),
                    request.form.get('respondent_phone', ''),
                    request.form.get('respondent_address', ''),
                    request.form.get('respondent_type', '自然人'),
                    request.form.get('respondent_cert_type', '身份证'),
                    request.form.get('respondent_uscc', ''),
                    request.form.get('respondent_gender', ''),
                    request.form.get('respondent_country', '中国'),
                    request.form.get('respondent_birth', ''),
                    request.form.get('respondent_age', '0'),
                    request.form.get('respondent_nation', '汉族'),
                    request.form.get('respondent_tel', ''),
                    request.form.get('respondent_residence', ''),
                    request.form.get('respondent_nature', ''),
                    request.form.get('respondent_legal_person', ''),
                    request.form.get('respondent_legal_title', ''),
                    request.form.get('respondent_reg_address', ''),
                    request.form.get('respondent_reg_country', '中国'),
                    request.form.get('preserve_amount', '0'),
                    request.form.get('court_name', ''),
                    request.form.get('guarantee_type', '提供保证人'),
                    request.form.get('guarantee_value', '0'),
                    request.form.get('guarantor_name', ''),
                    request.form.get('pledge_property_type', ''),
                    request.form.get('mortgage_property_type', ''),
                    request.form.get('guarantee_object', ''),
                    request.form.get('guarantee_remark', ''),
                    current_user
                )

                cur.execute(sql, values)
                case_id = cur.lastrowid
                conn.commit()
                flash(f'案件 {case_no} 创建成功', 'success')
                return redirect(url_for('view_case', case_id=case_id))
        except Exception as e:
            conn.rollback()
            flash(f'创建失败: {str(e)}', 'error')
            return redirect(url_for('add_case'))
        finally:
            conn.close()

    return render_template('case_form.html',
                          property_types=PROPERTY_TYPES,
                          guarantee_types=GUARANTEE_TYPES,
                          applicant_types=APPLICANT_TYPES,
                          respondent_types=RESPONDENT_TYPES)



@login_required
@app.route('/case/<int:case_id>/property/add', methods=['POST'])
def add_property(case_id):
    return redirect(url_for('view_case', case_id=case_id))

@app.route('/case/<int:case_id>/file/<int:file_id>/delete', methods=['POST'])
def delete_file(case_id, file_id):
    """删除案件文件"""
    conn = get_db()
    try:
        with conn.cursor() as cur:
            # 获取文件信息
            cur.execute("SELECT file_path FROM case_files WHERE id = %s AND case_id = %s", (file_id, case_id))
            file_row = cur.fetchone()
            if not file_row:
                flash('文件不存在', 'error')
                return redirect(url_for('view_case', case_id=case_id))
            
            file_path = file_row[0]
            
            # 删除物理文件
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except Exception as e:
                    print(f"删除物理文件失败: {e}")
            
            # 删除数据库记录
            cur.execute("DELETE FROM case_files WHERE id = %s", (file_id,))
            conn.commit()
            flash('文件删除成功', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'删除失败: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('view_case', case_id=case_id))

@login_required
@app.route('/case/<int:case_id>/edit', methods=['GET', 'POST'])
def edit_case(case_id):
    """保存案件编辑"""
    if request.method == 'GET':
        # 显示编辑页面
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM cases WHERE id = %s", (case_id,))
                case_row = cur.fetchone()
                if not case_row:
                    flash('案件不存在', 'error')
                    return redirect(url_for('index'))
                cur.execute("SHOW COLUMNS FROM cases")
                case_cols = [c[0] for c in cur.fetchall()]
                case = dict(zip(case_cols, case_row))
                
                cur.execute("SELECT * FROM property_clues WHERE case_id = %s", (case_id,))
                prop_rows = cur.fetchall()
                cur.execute("SHOW COLUMNS FROM property_clues")
                prop_cols = [c[0] for c in cur.fetchall()]
                properties = [dict(zip(prop_cols, row)) for row in prop_rows]
        finally:
            conn.close()
        return render_template('case_edit.html', case=case, properties=properties, guarantee_types=['提供保证人', '设定质押', '设定抵押', '交纳保证金'])
    
    conn = get_db()
    try:
        with conn.cursor() as cur:
            # 检查案件是否存在
            cur.execute("SELECT id FROM cases WHERE id = %s", (case_id,))
            if not cur.fetchone():
                flash('案件不存在', 'error')
                return redirect(url_for('index'))
            
            # 获取申请人类型，决定使用哪个字段
            applicant_type = request.form.get('applicant_type', '自然人')
            if applicant_type == '自然人':
                applicant_name = request.form.get('applicant_name', '')
                applicant_cert_no = request.form.get('applicant_cert_no', '')
                applicant_uscc = request.form.get('applicant_uscc', '')
                applicant_gender = request.form.get('applicant_gender', '')
            else:
                applicant_name = request.form.get('applicant_org_name', '')
                applicant_cert_no = ''
                applicant_uscc = request.form.get('applicant_org_cert_no', '')
                applicant_gender = ''
            
            respondent_type = request.form.get('respondent_type', '自然人')
            if respondent_type == '自然人':
                respondent_name = request.form.get('respondent_name', '')
                respondent_id = request.form.get('respondent_id', '')
                respondent_uscc = request.form.get('respondent_uscc', '')
                respondent_gender = request.form.get('respondent_gender', '')
            else:
                respondent_name = request.form.get('respondent_org_name', '')
                respondent_id = ''
                respondent_uscc = request.form.get('respondent_org_cert_no', '')
                respondent_gender = ''
            
            sql = """UPDATE cases SET
                case_name = %s, preserve_type = %s, case_reason = %s,
                applicant_name = %s, applicant_id = %s, applicant_phone = %s, 
                applicant_address = %s, applicant_type = %s,
                applicant_cert_type = %s, applicant_cert_no = %s, applicant_uscc = %s,
                applicant_gender = %s, applicant_country = %s, 
                applicant_birth = %s, applicant_age = %s, applicant_nation = %s,
                applicant_tel = %s, applicant_reg_address = %s, applicant_reg_country = %s,
                applicant_nature = %s, applicant_legal_person = %s, applicant_legal_title = %s,
                respondent_name = %s, respondent_id = %s, respondent_phone = %s,
                respondent_address = %s, respondent_type = %s,
                respondent_cert_type = %s, respondent_uscc = %s,
                respondent_gender = %s, respondent_country = %s,
                respondent_birth = %s, respondent_age = %s, respondent_nation = %s,
                respondent_tel = %s, respondent_residence = %s,
                respondent_nature = %s, respondent_legal_person = %s, respondent_legal_title = %s,
                respondent_reg_address = %s, respondent_reg_country = %s,
                preserve_amount = %s, court_name = %s,
                guarantee_type = %s, guarantee_value = %s,
                guarantor_name = %s, guarantee_property_type = %s, 
                pledge_property_type = %s, mortgage_property_type = %s,
                guarantee_object = %s, guarantee_remark = %s,
                updated_at = NOW()
                WHERE id = %s"""
            
            values = (
                request.form.get('case_name', ''),
                request.form.get('preserve_type', '诉讼保全'),
                request.form.get('case_reason', ''),
                applicant_name, applicant_cert_no,
                request.form.get('applicant_phone', ''),
                request.form.get('applicant_address', ''),
                applicant_type,
                request.form.get('applicant_cert_type', '身份证'),
                request.form.get('applicant_cert_no', ''),
                request.form.get('applicant_uscc', ''),
                applicant_gender,
                request.form.get('applicant_country', '中国'),
                request.form.get('applicant_birth', ''),
                request.form.get('applicant_age', ''),
                request.form.get('applicant_nation', '汉族'),
                request.form.get('applicant_tel', ''),
                request.form.get('applicant_reg_address', ''),
                request.form.get('applicant_reg_country', '中国'),
                request.form.get('applicant_nature', ''),
                request.form.get('applicant_legal_person', ''),
                request.form.get('applicant_legal_title', ''),
                respondent_name, respondent_id,
                request.form.get('respondent_phone', ''),
                request.form.get('respondent_address', ''),
                respondent_type,
                request.form.get('respondent_cert_type', '身份证'),
                request.form.get('respondent_uscc', ''),
                respondent_gender,
                request.form.get('respondent_country', '中国'),
                request.form.get('respondent_birth', ''),
                request.form.get('respondent_age', '0'),
                request.form.get('respondent_nation', '汉族'),
                request.form.get('respondent_tel', ''),
                request.form.get('respondent_residence', ''),
                request.form.get('respondent_nature', ''),
                request.form.get('respondent_legal_person', ''),
                request.form.get('respondent_legal_title', ''),
                request.form.get('respondent_reg_address', ''),
                request.form.get('respondent_reg_country', '中国'),
                request.form.get('preserve_amount', '0'),
                request.form.get('court_name', ''),
                request.form.get('guarantee_type', '提供保证人'),
                request.form.get('guarantee_value', '0'),
                request.form.get('guarantor_name', ''),
                request.form.get('guarantee_property_type', ''),
                request.form.get('pledge_property_type', ''),
                request.form.get('mortgage_property_type', ''),
                request.form.get('guarantee_object', ''),
                request.form.get('guarantee_remark', ''),
                case_id
            )
            
            cur.execute(sql, values)
            conn.commit()
            flash('案件更新成功', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'更新失败: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('view_case', case_id=case_id))


def add_property(case_id):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            sql = """INSERT INTO property_clues (
                case_id, property_type, owner, bank_name, bank_account,
                amount, currency, property_value, property_province,
                property_location, property_cert_no, property_detail_location,
                vehicle_brand_model, vehicle_plate_no, vehicle_location, vehicle_type,
                stock_reg_location, stock_company_name, stock_ratio,
                stock_name, stock_code, stock_quantity,
                fund_name, fund_quantity,
                bond_name, bond_face_value,
                equipment_name,
                property_city, property_area, property_area_unit,
                description
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
            values = (
                case_id, request.form.get('property_type', '存款'),
                request.form.get('owner', ''), request.form.get('bank_name', ''),
                request.form.get('bank_account', ''), request.form.get('amount', '0'),
                request.form.get('currency', 'CNY'), request.form.get('property_value', '0'),
                request.form.get('property_province', ''), request.form.get('property_location', ''),
                request.form.get('property_cert_no', ''), request.form.get('property_detail_location', ''),
                request.form.get('vehicle_brand_model', ''), request.form.get('vehicle_plate_no', ''),
                request.form.get('vehicle_location', ''), request.form.get('vehicle_type', ''),
                request.form.get('stock_reg_location', ''), request.form.get('stock_company_name', ''),
                request.form.get('stock_ratio') or None,
                request.form.get('stock_name', ''), request.form.get('stock_code', ''),
                request.form.get('stock_quantity', '0'),
                request.form.get('fund_name', ''), request.form.get('fund_quantity', '0'),
                request.form.get('bond_name', ''), request.form.get('bond_face_value', '0'),
                request.form.get('equipment_name', ''),
                request.form.get('property_city', ''), request.form.get('property_area', '0'),
                request.form.get('property_area_unit', '平方米'),
                request.form.get('description', '')
            )
            cur.execute(sql, values)
            conn.commit()
            flash('财产线索添加成功', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'添加失败: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('view_case', case_id=case_id))


@app.route('/case/<int:case_id>/property/<int:property_id>/edit', methods=['GET', 'POST'])
def edit_property(case_id, property_id):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            if request.method == 'GET':
                cur.execute("SELECT * FROM property_clues WHERE id = %s AND case_id = %s", (property_id, case_id))
                row = cur.fetchone()
                if not row:
                    flash('Property not found', 'error')
                    return redirect(url_for('view_case', case_id=case_id))
                cur.execute("SHOW COLUMNS FROM property_clues")
                cols = [c[0] for c in cur.fetchall()]
                prop = dict(zip(cols, row))
                return render_template('property_edit.html', prop=prop, case_id=case_id, property_id=property_id)
            sql = """UPDATE property_clues SET
                property_type = %s, owner = %s, bank_name = %s, bank_account = %s,
                amount = %s, currency = %s, property_value = %s, property_province = %s,
                property_location = %s, property_cert_no = %s, property_detail_location = %s,
                vehicle_brand_model = %s, vehicle_plate_no = %s, vehicle_location = %s, vehicle_type = %s,
                stock_reg_location = %s, stock_company_name = %s, stock_ratio = %s,
                stock_name = %s, stock_code = %s, stock_quantity = %s,
                fund_name = %s, fund_quantity = %s,
                bond_name = %s, bond_face_value = %s,
                equipment_name = %s,
                property_city = %s, property_area = %s, property_area_unit = %s,
                description = %s
            WHERE id = %s AND case_id = %s"""
            values = (
                request.form.get('property_type', '存款'),
                request.form.get('owner', ''),
                request.form.get('bank_name', ''),
                request.form.get('bank_account', ''),
                request.form.get('amount', '0'),
                request.form.get('currency', 'CNY'),
                request.form.get('property_value', '0'),
                request.form.get('property_province', ''),
                request.form.get('property_location', ''),
                request.form.get('property_cert_no', ''),
                request.form.get('property_detail_location', ''),
                request.form.get('vehicle_brand_model', ''),
                request.form.get('vehicle_plate_no', ''),
                request.form.get('vehicle_location', ''),
                request.form.get('vehicle_type', ''),
                request.form.get('stock_reg_location', ''),
                request.form.get('stock_company_name', ''),
                request.form.get('stock_ratio') or None,
                request.form.get('stock_name', ''),
                request.form.get('stock_code', ''),
                request.form.get('stock_quantity', '0'),
                request.form.get('fund_name', ''),
                request.form.get('fund_quantity', '0'),
                request.form.get('bond_name', ''),
                request.form.get('bond_face_value', '0'),
                request.form.get('equipment_name', ''),
                request.form.get('property_city', ''),
                request.form.get('property_area', '0'),
                request.form.get('property_area_unit', '平方米'),
                request.form.get('description', ''),
                property_id, case_id
            )
            cur.execute(sql, values)
            conn.commit()
            flash('财产线索更新成功', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'更新失败: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('view_case', case_id=case_id))


@app.route('/case/<int:case_id>/property/<int:property_id>/delete', methods=['POST'])
def delete_property(case_id, property_id):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM property_clues WHERE id = %s AND case_id = %s", (property_id, case_id))
            conn.commit()
            flash('财产线索删除成功', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'删除失败: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('view_case', case_id=case_id))


@app.route('/case/<int:case_id>/files/upload', methods=['POST'])
def upload_files(case_id):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT case_no, applicant_name, applicant_type, respondent_name, respondent_type, guarantee_type, guarantee_object FROM cases WHERE id = %s", (case_id,))
            row = cur.fetchone()
            if not row:
                flash('案件不存在', 'error')
                return redirect(url_for('index'))
            case_no = row[0]
            applicant_name = row[1]
            applicant_type = row[2]
            respondent_name = row[3]
            respondent_type = row[4]
            guarantee_type = row[5]
            guarantee_object = row[6]

            # 构建动态文件类别
            dynamic_categories = FILE_CATEGORIES.copy()
            dynamic_categories.append(f'申请人-{applicant_name}-{applicant_type}')
            dynamic_categories.append(f'被申请人-{respondent_name}-{respondent_type}')
            if guarantee_type == '提供保证人':
                dynamic_categories.append(f'提供保证人-{guarantee_object}')
            else:
                dynamic_categories.append(guarantee_type)

            uploaded_count = 0
            for category in dynamic_categories:
                if category in request.files:
                    files = request.files.getlist(category)
                    for file in files:
                        if file and file.filename:
                            filepath = save_uploaded_file(case_no, category, file)
                            if filepath:
                                cur.execute("""
                                    INSERT INTO case_files (case_id, file_category, file_name, file_path, upload_status)
                                    VALUES (%s, %s, %s, %s, 1)
                                """, (case_id, category, file.filename, filepath))
                                uploaded_count += 1
            conn.commit()
            flash(f'成功上传 {uploaded_count} 个文件', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'上传失败: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('view_case', case_id=case_id))


@login_required
@login_required
def edit_case(case_id):

    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT case_no FROM cases WHERE id = %s", (case_id,))
            row = cur.fetchone()
            if row:
                case_no = row[0]
                cur.execute("DELETE FROM case_files WHERE case_id = %s", (case_id,))
                cur.execute("DELETE FROM property_clues WHERE case_id = %s", (case_id,))
                cur.execute("DELETE FROM agents WHERE case_id = %s", (case_id,))
                cur.execute("DELETE FROM cases WHERE id = %s", (case_id,))
                conn.commit()
                case_dir = os.path.join(UPLOAD_BASE, case_no)
                if os.path.exists(case_dir):
                    shutil.rmtree(case_dir)
                flash('案件删除成功', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'删除失败: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('index'))


@app.route('/api/cases')
def api_cases():
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT c.id, c.case_no, c.case_name, c.applicant_name,
                       c.respondent_name, c.preserve_amount, c.court_name,
                       c.guarantee_type, c.status, c.created_at
                FROM cases c
                ORDER BY c.created_at DESC
            """)
            cases = []
            for row in cur.fetchall():
                cases.append({
                    'id': row[0], 'case_no': row[1], 'case_name': row[2],
                    'applicant_name': row[3], 'respondent_name': row[4],
                    'preserve_amount': str(row[5]), 'court_name': row[6],
                    'guarantee_type': row[7], 'status': row[8],
                    'created_at': str(row[9])
                })
            return jsonify({'success': True, 'data': cases})
    finally:
        conn.close()


# ==================== Excel模板下载与批量导入（增强版） ====================

def _build_case_data(row_data):
    """构建cases表插入数据"""
    return {
        'case_no': row_data.get('case_no', ''),
        'case_name': row_data.get('case_name', ''),
        'preserve_type': row_data.get('preserve_type', '诉讼保全'),
        'case_reason': row_data.get('case_reason', ''),
        'applicant_name': row_data.get('applicant_name', ''),
        'applicant_phone': row_data.get('applicant_phone', ''),
        'applicant_address': row_data.get('applicant_address', ''),
        'applicant_type': row_data.get('applicant_type', '自然人'),
        'applicant_cert_type': '身份证',
        'applicant_cert_no': row_data.get('applicant_cert_no', ''),
        'applicant_gender': row_data.get('applicant_gender', ''),
        'applicant_country': '中国',
        'applicant_birth': row_data.get('applicant_birth') or None,
        'applicant_age': row_data.get('applicant_age', '') or None,
        'applicant_nation': row_data.get('applicant_nation', '汉族'),
        'applicant_tel': row_data.get('applicant_tel', ''),
        'applicant_reg_address': row_data.get('applicant_reg_address', ''),
        'applicant_reg_country': '中国',
        'applicant_uscc': row_data.get('applicant_uscc', ''),
        'applicant_nature': row_data.get('applicant_nature', ''),
        'applicant_legal_person': row_data.get('applicant_legal_person', ''),
        'applicant_legal_title': row_data.get('applicant_legal_title', ''),
        'respondent_name': row_data.get('respondent_name', ''),
        'respondent_id': row_data.get('respondent_id', ''),
        'respondent_uscc': row_data.get('respondent_uscc', ''),
        'respondent_phone': row_data.get('respondent_phone', ''),
        'respondent_address': row_data.get('respondent_address', ''),
        'respondent_type': row_data.get('respondent_type', '自然人'),
        'respondent_cert_type': '身份证',
        'respondent_gender': row_data.get('respondent_gender', ''),
        'respondent_country': '中国',
        'respondent_birth': row_data.get('respondent_birth') or None,
        'respondent_age': row_data.get('respondent_age', '') or None,
        'respondent_nation': row_data.get('respondent_nation', '汉族'),
        'respondent_tel': row_data.get('respondent_tel', ''),
        'respondent_residence': row_data.get('respondent_residence', ''),
        'respondent_nature': row_data.get('respondent_nature', ''),
        'respondent_legal_person': row_data.get('respondent_legal_person', ''),
        'respondent_legal_title': row_data.get('respondent_legal_title', ''),
        'respondent_reg_address': row_data.get('respondent_reg_address', ''),
        'respondent_reg_country': '中国',
        'preserve_amount': row_data.get('preserve_amount', '0') or '0',
        'court_name': row_data.get('court_name', ''),
        'guarantee_type': row_data.get('guarantee_type', '提供保证人'),
        'guarantee_value': row_data.get('guarantee_value', '0') or '0',
        'guarantor_name': row_data.get('guarantor_name', ''),
        'pledge_property_type': row_data.get('guarantee_property_type', '') if row_data.get('guarantee_type') == '设定质押' else '',
        'mortgage_property_type': row_data.get('guarantee_property_type', '') if row_data.get('guarantee_type') == '设定抵押' else '',
        'guarantee_object': row_data.get('guarantee_object', ''),
        'guarantee_remark': row_data.get('guarantee_remark', ''),
        'delivery_address': row_data.get('delivery_address', ''),
        'contact_name': row_data.get('contact_name', ''),
        'contact_phone': row_data.get('contact_phone', ''),
        'remark': row_data.get('remark', '')
    }


def _build_property_data(case_id, row_data):
    """构建property_clues表插入数据"""
    ratio = row_data.get('stock_ratio', '')
    if ratio:
        ratio = re.sub(r'%$', '', ratio).strip()
        ratio = float(ratio) if ratio else None
    else:
        ratio = None
    
    # 辅助函数：空值返回None而不是'0'
    def val_or_none(key, default=None):
        v = row_data.get(key)
        if v is None or str(v).strip() == '' or str(v).strip() == 'nan':
            return default
        return v
    
    # 辅助函数：数值类型，空值返回None
    def num_or_none(key):
        v = row_data.get(key)
        if v is None or str(v).strip() == '' or str(v).strip() == 'nan':
            return None
        try:
            return float(v)
        except:
            return None
    
    return {
        'case_id': case_id,
        'property_type': row_data.get('property_type', '存款'),
        'owner': val_or_none('property_owner') or row_data.get('respondent_name', ''),
        'bank_name': val_or_none('bank_name', ''),
        'bank_account': val_or_none('bank_account', ''),
        'amount': num_or_none('amount') or num_or_none('preserve_amount') or 0,
        'currency': val_or_none('currency', 'CNY'),
        'property_value': num_or_none('property_value') or num_or_none('preserve_amount') or 0,
        'property_province': val_or_none('property_province', ''),
        'property_city': val_or_none('property_city', ''),
        'property_location': val_or_none('property_location', ''),
        'property_cert_no': val_or_none('property_cert_no', ''),
        'property_detail_location': val_or_none('property_detail_location', ''),
        'vehicle_plate_no': val_or_none('vehicle_plate_no', ''),
        'vehicle_location': val_or_none('vehicle_location', ''),
        'stock_name': val_or_none('stock_name', ''),
        'stock_code': val_or_none('stock_code', ''),
        'stock_quantity': num_or_none('stock_quantity'),
        'stock_reg_location': val_or_none('stock_reg_location', ''),
        'stock_company_name': val_or_none('stock_company_name', ''),
        'stock_ratio': ratio,
        'equipment_name': val_or_none('equipment_name', ''),
        'description': ''
    }


def _insert_case(cur, row_data, created_by=''):
    data = _build_case_data(row_data)
    data['created_by'] = created_by
    fields = list(data.keys())
    values = list(data.values())
    sql = f"INSERT INTO cases ({','.join(fields)}) VALUES ({','.join(['%s'] * len(fields))})"
    cur.execute(sql, values)
    return cur.lastrowid


def _insert_property(cur, case_id, row_data):
    data = _build_property_data(case_id, row_data)
    fields = list(data.keys())
    values = list(data.values())
    sql = f"INSERT INTO property_clues ({','.join(fields)}) VALUES ({','.join(['%s'] * len(fields))})"
    cur.execute(sql, values)


def _insert_agent(cur, case_id, row_data):
    if not row_data.get('agent_name'):
        return
    data = {
        'case_id': case_id,
        'agent_type': row_data.get('agent_type', '律师'),
        'agent_name': row_data.get('agent_name', ''),
        'agent_cert_no': row_data.get('agent_cert_no', ''),
        'agent_phone': row_data.get('agent_phone', ''),
        'agent_license_no': row_data.get('agent_license_no', ''),
        'agent_law_firm': row_data.get('agent_law_firm', '')
    }
    fields = list(data.keys())
    values = list(data.values())
    sql = f"INSERT INTO agents ({','.join(fields)}) VALUES ({','.join(['%s'] * len(fields))})"
    cur.execute(sql, values)


@app.route('/template/download')
def download_template():
    """下载Excel模板"""
    if not HAS_OPENPYXL:
        flash('请先安装 openpyxl: pip install openpyxl', 'error')
        return redirect(url_for('index'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "案件模板"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    required_fill = PatternFill(start_color="ffebee", end_color="ffebee", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [col[1] for col in EXCEL_TEMPLATE_COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # 示例1：主行
    example1 = [
        '保全2026001', '借款合同纠纷', '诉讼保全', '借款合同纠纷',
        '上海市静安区人民法院', '1000000', '北京市朝阳区XX路', '张三', '13800138000', '',
        '自然人', '张三', '110101199001011234', '男', '13800138000', '北京市朝阳区', '', '', '', '', '汉族', '',
        '自然人', '李四', '110101199002022345', '男', '13900139000', '北京市海淀区', '', '北京市海淀区', '', '', '', '',
        '存款', '李四', '中国工商银行', '6222021234567890123', '1000000', 'CNY', '1000000', '北京', '', '', '', '', '', '', '', '', '', '', '', '', '',
        '提供保证人', '100000', '王五', '', '存款', '',
        '律师', '赵六', '110101198001011111', '13600136000', '11101201901012345', '北京大成律师事务所'
    ]
    ws.append(example1)

    # 示例2：从行（追加财产）
    example2 = [''] * len(EXCEL_TEMPLATE_COLUMNS)
    example2[34] = '房产'
    example2[35] = '李四'
    example2[43] = '北京市海淀区XX小区'
    example2[44] = '京房权证海字第12345号'
    ws.append(example2)

    # 示例3：从行（追加被申请人）
    example3 = [''] * len(EXCEL_TEMPLATE_COLUMNS)
    example3[23] = '自然人'
    example3[24] = '王五'
    example3[25] = '110101199003033456'
    example3[27] = '13800138001'
    ws.append(example3)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for idx, col_def in enumerate(EXCEL_TEMPLATE_COLUMNS, 1):
        if col_def[2]:
            ws.cell(row=1, column=idx).fill = required_fill

    col_widths = [
        16, 20, 12, 16, 26, 12, 24, 14, 16, 16,
        12, 16, 22, 8, 16, 24, 16, 18, 12, 10, 10, 24,
        12, 16, 22, 8, 16, 24, 16, 24, 12, 16, 16, 24,
        12, 14, 18, 24, 12, 8, 12, 10, 10, 24, 18, 24, 14, 16, 12, 12, 12, 14, 16, 12, 12,
        12, 12, 14, 12, 14, 16, 12, 12, 14, 16, 18, 24
    ]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.freeze_panes = 'A2'

    ws_help = wb.create_sheet("填写说明")
    help_data = [
        ["字段", "说明", "可选值/示例"],
        ["案件编号*", "唯一标识，如：保全2026001", ""],
        ["案件名称*", "简要描述案件", ""],
        ["保全类别", "", "诉讼保全 / 诉前保全"],
        ["案由", "", "借款合同纠纷 / 买卖合同纠纷 ..."],
        ["申请法院*", "", "上海市静安区人民法院"],
        ["保全金额*", "数字，无需逗号", "1000000"],
        ["申请人类型", "", "自然人 / 法人 / 非法人组织"],
        ["申请人姓名*", "", ""],
        ["申请人身份证号*", "", ""],
        ["申请人手机号*", "", ""],
        ["被申请人类型", "", "自然人 / 法人 / 非法人组织"],
        ["被申请人姓名*", "", ""],
        ["被申请人身份证号*", "", ""],
        ["财产类型*", "", "存款 / 房产 / 股票 / 股权 / 交通运输工具 / 机器设备 / 货币现金 / 其他"],
        ["币种", "", "CNY / USD / EUR"],
        ["担保方式", "", "提供保证人 / 设定抵押 / 设定质押 / 交纳保证金"],
        ["申请人性别", "", "男 / 女"],
        ["被申请人性别", "", "男 / 女"],
        ["财产所在省份", "", "北京 / 上海 / 广东 / 浙江 ..."],
        ["", "", ""],
        ["多行说明", "同一案件支持多行：", ""],
        ["", "1. 主行：案件编号非空，填写完整案件信息", ""],
        ["", "2. 从行：案件编号留空，仅填写财产/被申请人/代理人信息，归属上一主行", ""],
        ["", "3. 一个案件可有多条从行，用于添加多个财产或多个被申请人", ""],
    ]
    for row in help_data:
        ws_help.append(row)
    for cell in ws_help[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws_help.column_dimensions['A'].width = 20
    ws_help.column_dimensions['B'].width = 60
    ws_help.column_dimensions['C'].width = 50

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='案件批量导入模板.xlsx'
    )


@login_required
@app.route('/template/upload', methods=['POST'])
def upload_template():
    """上传Excel批量创建案件（支持多行模式）"""
    if not HAS_OPENPYXL:
        flash('请先安装 openpyxl: pip install openpyxl', 'error')
        return redirect(url_for('index'))

    if 'excel_file' not in request.files:
        flash('请选择Excel文件', 'error')
        return redirect(url_for('index'))

    file = request.files['excel_file']
    if not file or not file.filename:
        flash('请选择Excel文件', 'error')
        return redirect(url_for('index'))

    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('请上传 .xlsx 或 .xls 格式的Excel文件', 'error')
        return redirect(url_for('index'))

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        field_map = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            header_clean = str(header).replace('*', '').strip()
            for col_def in EXCEL_TEMPLATE_COLUMNS:
                if col_def[1].replace('*', '') == header_clean:
                    field_map[idx] = col_def[0]
                    break

        if not field_map:
            flash('无法识别Excel表头，请使用系统提供的模板', 'error')
            return redirect(url_for('index'))

        # 获取当前登录用户
        current_user = ''
        try:
            conn2 = get_db()
            with conn2.cursor() as cur2:
                cur2.execute("SELECT config_value FROM system_config WHERE config_key = 'login_username'")
                row = cur2.fetchone()
                if row:
                    current_user = row[0]
            conn2.close()
        except:
            pass
        
        conn = get_db()
        created_cases = []
        failed_cases = []

        try:
            with conn.cursor() as cur:
                current_case_id = None
                current_case_no = None

                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    has_data = any(v is not None and str(v).strip() != '' for v in row)
                    if not has_data:
                        continue

                    row_data = {}
                    for idx, val in enumerate(row):
                        if idx in field_map and val is not None:
                            row_data[field_map[idx]] = str(val).strip()

                    case_no = row_data.get('case_no', '')

                    if case_no:
                        cur.execute("SELECT id FROM cases WHERE case_no = %s", (case_no,))
                        if cur.fetchone():
                            failed_cases.append(f'{case_no} (第{row_idx}行): 案件编号已存在')
                            current_case_id = None
                            current_case_no = None
                            continue

                        required_fields = ['case_name', 'applicant_name', 'applicant_cert_no',
                                           'respondent_name', 'respondent_id', 'court_name',
                                           'preserve_amount', 'property_type']
                        missing = [f for f in required_fields if not row_data.get(f)]
                        if missing:
                            failed_cases.append(f'{case_no} (第{row_idx}行): 缺少必填字段 {", ".join(missing)}')
                            current_case_id = None
                            current_case_no = None
                            continue

                        current_case_id = _insert_case(cur, row_data, current_user)
                        current_case_no = case_no
                        created_cases.append(case_no)

                        _insert_property(cur, current_case_id, row_data)
                        _insert_agent(cur, current_case_id, row_data)

                    else:
                        if current_case_id is None:
                            failed_cases.append(f'第{row_idx}行: 案件编号为空且无前序主行，无法识别归属')
                            continue

                        if row_data.get('property_type'):
                            _insert_property(cur, current_case_id, row_data)

                        if row_data.get('agent_name'):
                            _insert_agent(cur, current_case_id, row_data)

                conn.commit()
        except Exception as e:
            conn.rollback()
            flash(f'批量导入失败: {str(e)}', 'error')
            return redirect(url_for('index'))
        finally:
            conn.close()

        msg_parts = []
        if created_cases:
            msg_parts.append(f'成功创建 {len(created_cases)} 个案件')
        if failed_cases:
            msg_parts.append(f'失败 {len(failed_cases)} 个')
            for f in failed_cases[:10]:
                flash(f, 'error')

        flash(' / '.join(msg_parts), 'success' if created_cases else 'error')

    except Exception as e:
        flash(f'Excel解析失败: {str(e)}', 'error')

    return redirect(url_for('index'))




# ==================== 批量案件文件上传（zip压缩包） ====================

@app.route('/batch/files/upload', methods=['POST'])
@app.route('/batch/files/upload', methods=['POST'])
def batch_files_upload():
    """批量上传案件文件（zip压缩包）

    压缩包结构要求：
    保全2026001/
        保全申请书/01_保全申请书.pdf
        起诉状/起诉状.doc
        身份证明材料/03_申请人证照.pdf
        担保材料/06_担保函.pdf
    保全2026002/
        保全申请书/01_保全申请书.pdf
        ...
    """
    if 'zip_file' not in request.files:
        flash('请选择zip压缩包', 'error')
        return redirect(url_for('index'))

    file = request.files['zip_file']
    if not file or not file.filename:
        flash('请选择zip压缩包', 'error')
        return redirect(url_for('index'))

    if not file.filename.endswith('.zip'):
        flash('请上传 .zip 格式的压缩包', 'error')
        return redirect(url_for('index'))

    import zipfile
    import tempfile

    results = []
    total_files = 0
    skipped_files = []

    temp_dir = None
    try:
        temp_dir = tempfile.mkdtemp(prefix='batch_upload_')
        zip_path = os.path.join(temp_dir, 'upload.zip')
        file.save(zip_path)

        extract_dir = os.path.join(temp_dir, 'extracted')
        os.makedirs(extract_dir, exist_ok=True)

        with zipfile.ZipFile(zip_path, 'r') as zf:
            zf.extractall(extract_dir)

        conn = get_db()
        try:
            with conn.cursor() as cur:
                for case_folder in os.listdir(extract_dir):
                    case_path = os.path.join(extract_dir, case_folder)
                    if not os.path.isdir(case_path):
                        continue

                    cur.execute("SELECT id, case_no FROM cases WHERE case_no = %s", (case_folder,))
                    case_row = cur.fetchone()
                    if not case_row:
                        results.append({'case_no': case_folder, 'status': '跳过', 'reason': '案件不存在'})
                        continue

                    case_id, case_no = case_row
                    case_files_count = 0

                    for category in os.listdir(case_path):
                        cat_path = os.path.join(case_path, category)
                        if not os.path.isdir(cat_path):
                            continue

                        # 获取案件信息用于构建动态类别
                        cur.execute("SELECT applicant_name, applicant_type, respondent_name, respondent_type, guarantee_type, guarantee_object FROM cases WHERE id = %s", (case_id,))
                        case_info = cur.fetchone()
                        
                        # 构建动态文件类别列表
                        dynamic_categories = FILE_CATEGORIES.copy()
                        if case_info:
                            dynamic_categories.append(f'申请人-{case_info[0]}-{case_info[1]}')
                            dynamic_categories.append(f'被申请人-{case_info[2]}-{case_info[3]}')
                            if case_info[4] == '提供保证人':
                                dynamic_categories.append(f'提供保证人-{case_info[5]}')
                            else:
                                dynamic_categories.append(case_info[4])
                        
                        # 同时支持简化的类别名称
                        dynamic_categories.append('申请人')
                        dynamic_categories.append('被申请人')

                        if category not in dynamic_categories:
                            skipped_files.append(f"{case_no}/{category}/ (未知类别)")
                            continue

                        for filename in os.listdir(cat_path):
                            src_path = os.path.join(cat_path, filename)
                            if not os.path.isfile(src_path):
                                continue

                            # 校验扩展名
                            if not allowed_file(filename):
                                skipped_files.append(f"{case_no}/{category}/{filename} (不允许的文件类型)")
                                continue

                            # 走标准保存逻辑（secure_filename + 覆盖处理）
                            from werkzeug.datastructures import FileStorage
                            with open(src_path, 'rb') as fsrc:
                                fs = FileStorage(
                                    stream=io.BytesIO(fsrc.read()),
                                    filename=filename,
                                    name=filename
                                )
                                filepath = save_uploaded_file(case_no, category, fs, applicant_name=case_info[0] if case_info else None, respondent_name=case_info[2] if case_info else None)

                            if filepath:
                                # 使用新的文件名（案号-案件材料目录.扩展名）
                                new_filename = os.path.basename(filepath)
                                cur.execute("""
                                    INSERT INTO case_files (case_id, file_category, file_name, file_path, upload_status)
                                    VALUES (%s, %s, %s, %s, 1)
                                """, (case_id, category, new_filename, filepath))
                                case_files_count += 1
                                total_files += 1
                            else:
                                skipped_files.append(f"{case_no}/{category}/{filename} (保存失败)")

                    results.append({'case_no': case_no, 'status': '成功', 'count': case_files_count})

                conn.commit()
        except Exception as e:
            conn.rollback()
            flash(f'批量上传失败: {str(e)}', 'error')
            return redirect(url_for('index'))
        finally:
            conn.close()

        success_cases = [r for r in results if r['status'] == '成功']
        skip_cases = [r for r in results if r['status'] == '跳过']

        msg = f'批量上传完成：成功 {len(success_cases)} 个案件，共 {total_files} 个文件'
        if skip_cases:
            msg += f'，跳过 {len(skip_cases)} 个案件（不存在）'
        flash(msg, 'success')

        for r in skip_cases[:5]:
            flash(f"跳过 {r['case_no']}: {r['reason']}", 'warning')

        if skipped_files:
            for sf in skipped_files[:5]:
                flash(f"跳过文件: {sf}", 'warning')
            if len(skipped_files) > 5:
                flash(f"... 还有 {len(skipped_files)-5} 个文件被跳过", 'warning')

    except Exception as e:
        flash(f'压缩包处理失败: {str(e)}', 'error')
    finally:
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)

    return redirect(url_for('index'))


@login_required
@app.route('/auto-filing/start', methods=['POST'])
def auto_filing_start():
    """启动全部批量自动立案"""
    import subprocess
    import threading
    
    def run_filing():
        global auto_filing_status
        auto_filing_status['running'] = True
        auto_filing_status['start_time'] = datetime.now()
        auto_filing_status['total'] = 0
        auto_filing_status['current'] = 0
        auto_filing_status['current_case'] = None
        auto_filing_status['completed'] = []
        auto_filing_status['failed'] = []
        try:
            proc = subprocess.Popen(
                cwd=r'/app',
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                creationflags=subprocess.CREATE_NEW_CONSOLE
            )
            stdout, stderr = proc.communicate(timeout=7200)
            app.logger.info('batch_auto_filing.py 完成，返回码: %d', proc.returncode)
            if stdout:
                app.logger.info('stdout: %s', stdout[-2000:])
            if stderr:
                app.logger.error('stderr: %s', stderr[-2000:])
        except subprocess.TimeoutExpired:
            app.logger.error('batch_auto_filing.py 超时')
            proc.kill()
        except Exception as e:
            app.logger.error('batch_auto_filing.py 异常: %s', e)
        finally:
            auto_filing_status['running'] = False
            auto_filing_status['current_case'] = None
            try:
                import json
                with open(r'/app//auto_filing_progress.json', 'r', encoding='utf-8') as f:
                    progress = json.load(f)
                progress['running'] = False
                progress['current_case'] = None
                with open(r'/app//auto_filing_progress.json', 'w', encoding='utf-8') as f:
                    json.dump(progress, f, ensure_ascii=False)
            except Exception:
                pass
    
    thread = threading.Thread(target=run_filing)
    thread.daemon = True
    thread.start()
    
    flash('批量自动立案已启动', 'success')
    return redirect(url_for('auto_filing_status_page'))


@login_required
@app.route('/auto-filing/selected', methods=['POST'])
def auto_filing_selected():
    """对选中的案件自动立案"""
    import subprocess
    import threading
    
    case_nos = request.form.getlist('case_nos')
    app.logger.info('自动立案接收案件数量: %s, 案件列表: %s' % (len(case_nos), case_nos))
    # 写入标记文件方便查看
    debug_info = "时间: %s/n案件数量: %s/n案件列表: %s/n" % (datetime.now(), len(case_nos), case_nos)
    with open(r'/app/auto_filing_debug.txt', 'w', encoding='utf-8') as f:
        f.write(debug_info)
    if not case_nos:
        flash('未选择任何案件', 'warning')
        return redirect(url_for('index'))
    
    def run_selected(cases):
        import time
        global auto_filing_status
        auto_filing_status['running'] = True
        auto_filing_status['total'] = len(cases)
        auto_filing_status['current'] = 0
        auto_filing_status['current_case'] = None
        auto_filing_status['completed'] = []
        auto_filing_status['failed'] = []
        auto_filing_status['start_time'] = datetime.now()
        app.logger.info(f'run_selected开始，共{len(cases)}个案件')
        for idx, case_no in enumerate(cases, 1):
            auto_filing_status['current'] = idx
            auto_filing_status['current_case'] = case_no
            app.logger.info(f'开始立案第 {idx}/{len(cases)} 个案件: {case_no}')
            try:
                result = subprocess.run(
                    ['python', 'final_auto_upload_db_v3.py', case_no],
                    cwd=r'/app',
                    capture_output=True,
                    text=True,
                    timeout=600
                )
                app.logger.info(f'案件 {case_no} 返回码: {result.returncode}')
                if result.returncode != 0:
                    app.logger.error(f'案件 {case_no} 失败: {result.stderr[:500]}')
                    auto_filing_status['failed'].append(case_no)
                    continue
            except subprocess.TimeoutExpired:
                app.logger.error(f'案件 {case_no} 超时')
                auto_filing_status['failed'].append(case_no)
                continue
            except Exception as e:
                app.logger.error(f'案件 {case_no} 异常: {e}')
                auto_filing_status['failed'].append(case_no)
                continue
            app.logger.info(f'案件 {case_no} 开始更新状态')
            conn = get_db()
            try:
                with conn.cursor() as cur:
                    cur.execute("UPDATE cases SET status = 1 WHERE case_no = %s", (case_no,))
                    conn.commit()
                    app.logger.info(f'案件 {case_no} 状态已更新为已立案')
                    auto_filing_status['completed'].append(case_no)
            except Exception as e:
                app.logger.error(f'案件 {case_no} 更新状态失败: {e}')
                auto_filing_status['failed'].append(case_no)
            finally:
                conn.close()
            app.logger.info(f'案件 {case_no} 休眠5秒后继续')
            time.sleep(5)
            app.logger.info(f'案件 {case_no} 休眠结束，继续下一个')
        auto_filing_status['running'] = False
        auto_filing_status['current_case'] = None
        app.logger.info('所有案件处理完成')
    
    thread = threading.Thread(target=run_selected, args=(case_nos,))
    # thread.daemon = True
    thread.start()
    
    flash('已启动 %s 个案件的自动立案' % len(case_nos), 'success')
    return redirect(url_for('index'))


@login_required
@app.route('/auto-filing/stop', methods=['POST'])
def auto_filing_stop():
    """停止自动立案"""
    import os
    stop_file = r'/app//auto_filing_stop.flag'
    with open(stop_file, 'w', encoding='utf-8') as f:
        f.write('stop')
    flash('已发送停止指令，当前案件处理完后将停止', 'warning')
    return redirect(url_for('auto_filing_status_page'))


def read_filing_progress():
    import json
    try:
        with open(r'/app//auto_filing_progress.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return auto_filing_status

@login_required
@app.route('/auto-filing/status')
def auto_filing_status_page():
    """查看自动立案进度"""
    return render_template('auto_filing_status.html', status=read_filing_progress())


@login_required
@app.route('/api/auto-filing/status')
def api_auto_filing_status():
    """API：获取自动立案状态"""
    return jsonify(read_filing_progress())


@app.route('/sync/filing-status', methods=['POST'])
@login_required
def sync_filing_status_api():
    """同步立案状态(后台异步执行)"""
    import subprocess, json, os, time
    try:
        # 使用后台进程执行同步，避免HTTP超时
        log_file = r'C:\court-auto-filing\sync_bg.log'
        with open(log_file, 'w', encoding='utf-8') as f:
            f.write(f"同步任务启动于 {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        subprocess.Popen(
            [r'C:\Users\Administrator\AppData\Local\Programs\Python\Python312\python.exe', r'C:\court-auto-filing\sync_filing_status.py'],
            cwd=r'C:\court-auto-filing',
            stdout=open(log_file, 'a', encoding='utf-8'),
            stderr=subprocess.STDOUT,
            creationflags=subprocess.CREATE_NEW_CONSOLE if os.name == 'nt' else 0
        )
        
        return json.dumps({'success': True, 'message': '同步任务已启动，请在几分钟后刷新页面查看结果'}), 200, {'Content-Type': 'application/json'}
    except Exception as e:
        return json.dumps({'success': False, 'message': str(e)}), 500, {'Content-Type': 'application/json'}



@app.route('/sync/clear-new', methods=['POST'])
@login_required
def clear_new_mark():
    """清除新增标记"""
    import pymysql
    try:
        conn = pymysql.connect(**STATUS_DB_CONFIG)
        with conn.cursor() as cur:
            tables = ['filing_status_trial', 'filing_status_execution', 'filing_status_preservation',
                      'filing_status_mediation', 'filing_status_bankruptcy', 'filing_status_petition']
            for table in tables:
                try:
                    cur.execute(f"UPDATE {table} SET is_new = 0")
                except:
                    pass
            conn.commit()
        conn.close()
        return json.dumps({'success': True, 'message': '已清除新增标记'}), 200, {'Content-Type': 'application/json'}
    except Exception as e:
        return json.dumps({'success': False, 'message': str(e)}), 500, {'Content-Type': 'application/json'}



if __name__ == '__main__':
    os.makedirs(UPLOAD_BASE, exist_ok=True)
    print("启动后台管理系统...")
    print("访问地址: http://localhost:5000")
    app.run(host='0.0.0.0', port=5000, debug=False)



@app.route('/sync/status/history')
@login_required
def sync_status_history():
    """查看同步历史"""
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT sync_time, total_cases, updated_cases, error_msg 
                FROM court_filing_status.sync_log 
                ORDER BY sync_time ASC LIMIT 50
            """)
            logs = cur.fetchall()
        return render_template('sync_history.html', logs=logs)
    except Exception as e:
        flash('查询失败: ' + str(e), 'error')
        return redirect(url_for('index'))
    finally:
        conn.close()
